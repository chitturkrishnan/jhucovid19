#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Apr 06 11:25:56 2020

@author: kchittur
"""

# a hack to examine and plot git hub hopkins data on sars cov2
# creates a bunch of different plots - when I can, I will add some
# documentation
# collect libraries you will need - for reading csv files, plotting etc

import matplotlib.pyplot as plt
from matplotlib import rc
import matplotlib
#rc('font',**{'family':'serif','serif':['Palatino']})
rc('text', usetex=True)
matplotlib.rcParams['text.latex.preamble'] = [r'\boldmath']
from matplotlib import rcParams
rcParams.update({'figure.autolayout': True})
#import sys
import numpy as np
from csv import reader
from csv import *
import csv, glob,natsort
import os, glob, sys, cv2

# pulled in data from the JHU github 
# all directories are with respect to $HOME/COVID-19 
# adjust accordingly

populationdata = 'populationdata/'
worldpopulation = populationdata+'WPP2019_POP_F01_1_TOTAL_POPULATION_BOTH_SEXES.xlsx'
timeseries = 'csse_covid_19_data/csse_covid_19_time_series/'
confirmedGlobal = timeseries+'time_series_covid19_confirmed_global.csv'
deathsGlobal = timeseries+'time_series_covid19_deaths_global.csv'
recoveredGlobal = timeseries+'time_series_covid19_recovered_global.csv'
confirmedUS = timeseries+'time_series_covid19_confirmed_US.csv'
deathsUS = timeseries+'time_series_covid19_deaths_US.csv'
recoveredUS = timeseries+'time_series_covid19_recovered_US.csv'

def getlistofcountries(filename):
    csvfile = open(filename,'r')
    mycsvreader = csv.reader(csvfile,delimiter=',')
    therows = []
    for row in mycsvreader:
        therows.append(row)
    for i in range(len(therows[0])):
        thiscell = therows[0][i]
        if (thiscell == "Country/Region"):
            cc = i
            break
    thecountriesX = []

    for i in range(1,len(therows)):
        thecountriesX.append(therows[i][cc])
    countrynames = list(set(thecountriesX))
    return natsort.natsorted(countrynames)

from usstates import *
statenames = []
for key, value in states.items():
    statenames.append(str(value))

from dateutil.parser import parse
def is_date(string, fuzzy=False):
    """
    Return whether the string can be interpreted as a date.

    :param string: str, string to check for date
    :param fuzzy: bool, ignore unknown tokens in string if True
    """
    try: 
        parse(string, fuzzy=fuzzy)
        return True
    except ValueError:
        return False


# get the date from the files - read the first line really

def getstatedates(filename):
    csvfile = open(filename,'r')
    mycsvreader = csv.reader(csvfile,delimiter=',')
    thedates = []
    therows = []
    for row in mycsvreader:
        therows.append(row)
# examine the first row to find out where the date starts
    for i in range(len(therows[0])):
        thiscell = therows[0][i]
        if (is_date(thiscell)):
            cn = i 
            break
    for j in range(cn,len(therows[0])):
        thedates.append(therows[0][j])
    return thedates

def getstatedata(filename,thecountry):
#    print ("Working with ", filename, " and for ", thecountry)
    csvfile = open(filename,'r')
    mycsvreader = csv.reader(csvfile,delimiter=',')
    therows = []
    for row in mycsvreader:
        therows.append(row)
    rown = []
    nrows = 0

    for i in range(len(therows)):
        thiscountry = therows[i][6]
        if (thiscountry == thecountry):
#            print (thiscountry)
            rown.append(i)
            nrows = nrows + 1
    nl = len(therows[0])
#    print ("nl is ", nl)
    for i in range(len(therows[0])):
        thiscell = therows[0][i]
        if (is_date(thiscell)):
            cn = i
            break      
    thesum = []
#    print ("cn is ", cn)
#    print (rown, len(rown))
    for i in range(cn,nl):
        sum = 0.0
        for j in range(len(rown)):
            thisr = int(rown[j])
#            print ("i ",i, "j ",j, "thisr ", thisr, "therows thisr i", therows[thisr][i])
            sum = sum + float(therows[thisr][i])

        thesum.append(sum)          
    return thesum

def getdates(filename):
    print ("reading ", filename)
    csvfile = open(filename,'r')
    mycsvreader = csv.reader(csvfile,delimiter=',')
    thedates = []
    therows = []
    for row in mycsvreader:
        therows.append(row)
    for i in range(len(therows[0])):
        thiscell = therows[0][i]
        if (is_date(thiscell)):
            cn = i 
            break
    for j in range(cn,len(therows[0])):
        thedates.append(therows[0][j])
    return thedates

def getcountrydata(filename,thecountry):
    csvfile = open(filename,'r')
    mycsvreader = csv.reader(csvfile,delimiter=',')
    therows = []
    thenumbers= []
    for row in mycsvreader:
        therows.append(row)
    rown = []
    nrows = 0
    for i in range(len(therows)):
        thiscountry = therows[i][1]
        if (thiscountry == thecountry):
            rown.append(i)
            nrows = nrows + 1
    nl = len(therows[0])
    for i in range(len(therows[0])):
        thiscell = therows[0][i]
        if (is_date(thiscell)):
            cn = i
            break      
    thesum = []
    for i in range(cn,nl):
        sum = 0
        for j in range(len(rown)):
            thisr = int(rown[j])
            sum = sum + int(therows[thisr][i])
        thesum.append(sum)          
    return thesum

# this creates a text slide of the same size as the plots -
# you can use this if you create a video from a bunch of png files
# where you can provide an explanation of the slides the user may see

def createtextslide(thetexts,xpos,ypos,outputname):
    import matplotlib.pyplot as plt
    from matplotlib import rc
    import matplotlib
#    rc('font',**{'family':'serif','serif':['Palatino']})
    rc('text', usetex=True)
    from matplotlib import rcParams
    rcParams.update({'figure.autolayout': True})
    rcParams['figure.dpi'] = 200
    import numpy as np
    plt.tick_params(
            axis='both',          # changes apply to the x-axis
            which='both',      # both major and minor ticks are affected
            bottom=False,      # ticks along the bottom edge are off
            top=False,         # ticks along the top edge are off
            labelbottom=False, # labels along the bottom edge are off
            labelleft=False)

    xdata = np.linspace(0,10,11)
    ydata = np.linspace(0,10,11)
    plt.plot(xdata,ydata,linestyle='None')
    for i in range(len(thetexts)):
        thetext = thetexts[i]
        thex = xpos[i]
        they = ypos[i]
        plt.text(thex,they,thetext,fontsize=14)
    plt.savefig(outputname)
    plt.clf()

# this takes a series of images (png files) and can create a 
# video file - choose fps to be small if you want the images to stay
# longer on the screen - fps is frames per second
# you can create a avi or a mp4 file 

def createvideofile(filenames,fps,fileformat,outputname):
    import cv2
    img_array = []
    print ("Assembling png files into an mp4 file")
    for i in range(len(filenames)):
        filename = filenames[i]
        print ("working on ", filename)
        img = cv2.imread(filename)
        height, width, layers = img.shape
        size = (width,height)
    
        img_array.append(img)
    out = cv2.VideoWriter(outputname,cv2.VideoWriter_fourcc(*fileformat),fps, size)
    for i in range(len(img_array)):
        out.write(img_array[i])
    out.release()
    print ("Created ", outputname)

def createUSstateratio(filenames,filetitles,ndays,countryname,prefix):
    import matplotlib.ticker as ticker
    import matplotlib.pyplot as plt
    print ("Creating ratios for" , countryname)
#    print ("Inside this one")
    country = countryname
    ax = plt.axes()
    nxmajor = int(ndays/20)
    nxminor = int(ndays/10)
    ax.xaxis.set_major_locator(ticker.MultipleLocator(nxmajor))
    ax.xaxis.set_minor_locator(ticker.MultipleLocator(nxminor))
    plt.xticks(rotation=90,fontsize=10,fontweight='bold')
    plt.xlabel("Date")
    plt.ylabel("Ratio")             
    one = str(filetitles[0])
    two = str(filetitles[1])
    plt.title(r" %s over %s %s days for %s " %(one,two,ndays,country),fontsize=12)            
    plt.tight_layout
    plt.grid()
    filename = filenames[0]
    xdata = getstatedates(filename)
#    print (xdata)
#    print (xdata)
    ydata1 = getstatedata(filename,country)
#    print (ydata1)
    filename = filenames[1]
    ydata2 = getstatedata(filename,country)    
#    print (xdata)
#    print (ydata1)
#    print (ydata2)
    theratio = []
    for i in range(len(ydata1)-1):
        if (ydata1[i] == 0 or ydata2[i] == 0):
            theratio.append(0)
        else:
            theratio.append(ydata1[i]/ydata2[i])      
    thislabel=country
    plt.plot(xdata[-ndays+1:],theratio[-ndays+1:],label='ratio')
    outputname = prefix+filetitles[0]+filetitles[1]+'ratiosfor'+country+'.png'
    plt.savefig(outputname)
#    print (countryname,theratio)
#    print ("Done here")
    plt.clf()

def createglobalratio(filenames,filetitles,ndays,countryname,prefix):
    import matplotlib.ticker as ticker
    import matplotlib.pyplot as plt
    from matplotlib import rcParams
    rcParams.update({'figure.autolayout': True})
    rcParams['figure.dpi'] = 200
    print (filetitles[0],'/',filetitles[1], ' for ', countryname)
    country = countryname
    ax = plt.axes()
    nxmajor = int(ndays/20)
    nxminor = int(ndays/10)
    ax.xaxis.set_major_locator(ticker.MultipleLocator(nxmajor))
    ax.xaxis.set_minor_locator(ticker.MultipleLocator(nxminor))
    plt.xticks(rotation=90,fontsize=10,fontweight='bold')
    plt.xlabel("Date")
    plt.ylabel("Ratio")             
    one = str(filetitles[0])
    two = str(filetitles[1])
    plt.title(r"$\frac{%s}{%s}$ over %s days for %s " %(one,two,ndays,country),fontsize=12)            
    plt.tight_layout
    plt.grid()
    filename = filenames[0]
    xdata = getdates(filename)
    ydata1 = getcountrydata(filename,country)
    filename = filenames[1]
    ydata2 = getcountrydata(filename,country)    
    theratio = []
    for i in range(len(ydata1)):
        if (ydata1[i] == 0 or ydata2[i] == 0):
            theratio.append(0)
        else:
            theratio.append(ydata1[i]/ydata2[i])      
    thislabel=country
    plt.plot(xdata[-ndays+1:],theratio[-ndays+1:],label='ratio')
    outputname = prefix+filetitles[0]+filetitles[1]+'ratiosfor'+country+'.png'
    plt.savefig(outputname)
    print ("Created ", outputname)
    plt.clf()


def createglobalratioX(filenames,filetitles,ndays,countryname,prefix):
    import matplotlib.ticker as ticker
    import matplotlib.pyplot as plt
    from matplotlib import rcParams
    rcParams.update({'figure.autolayout': True})
    rcParams['figure.dpi'] = 200
    print (filetitles[0],'/',filetitles[1], ' for ', countryname)
    country = countryname
    ax = plt.axes()
    nxmajor = int(ndays/20)
    nxminor = int(ndays/10)
    ax.xaxis.set_major_locator(ticker.MultipleLocator(nxmajor))
    ax.xaxis.set_minor_locator(ticker.MultipleLocator(nxminor))
    plt.xticks(rotation=90,fontsize=10,fontweight='bold')
    plt.xlabel("Date")
    plt.ylabel("Ratio")             
    one = str(filetitles[0])
    two = str(filetitles[1])
    plt.title(r"$\frac{%s}{%s}$ over %s days for %s " %(one,two,ndays,country),fontsize=12)            
    plt.tight_layout
    plt.grid()
    filename = filenames[0]
    xdata = getdates(filename)
    ydata1 = getcountrydata(filename,country)
    filename = filenames[1]
    ydata2 = getcountrydata(filename,country)    
    theratio = []
    for i in range(len(ydata1)):
        if (ydata1[i] == 0 or ydata2[i] == 0):
            theratio.append(0)
        else:
            theratio.append(ydata1[i]/ydata2[i])      
    thislabel=country
    rmax = np.max(theratio)
    rindex = xdata[theratio.index(rmax)] 

    thelabel='ratio - Max on '+str(rindex)
    plt.plot(xdata[-ndays+1:],theratio[-ndays+1:],label=thelabel)
    plt.axvline(x=rindex,ls=':',color='r')
    plt.legend(loc='best')
    outputname = prefix+filetitles[0]+filetitles[1]+'ratiosfor'+country+'.png'
    plt.savefig(outputname)
    print ("Created ", outputname)
    plt.clf()

def getnratio(x):
    import numpy as np
    xmax = np.max(x)
    xn = []
    for i in range(len(x)):
        xn.append(x[i]/xmax)
    return xn
        

def createallglobalratioX(filenames,filetitles,ndays,countrynames):
    import matplotlib.ticker as ticker
    import matplotlib.pyplot as plt
    from matplotlib import rcParams
    from openpyxl import Workbook
    outwb = Workbook()
    outsheet = outwb.create_sheet(title='Maximum')
    outsheet.cell(row=1,column=1).value = 'Country'
    outsheet.cell(row=1,column=2).value = 'Date'
    rcParams.update({'figure.autolayout': True})
    rcParams['figure.dpi'] = 200
    ax = plt.axes()
    nxmajor = int(ndays/20)
    nxminor = int(ndays/10)
    ax.xaxis.set_major_locator(ticker.MultipleLocator(nxmajor))
    ax.xaxis.set_minor_locator(ticker.MultipleLocator(nxminor))
    plt.xticks(rotation=90,fontsize=10,fontweight='bold')
    plt.xlabel("Date")
    plt.ylabel("Ratio")             
    one = str(filetitles[0])
    two = str(filetitles[1])
    plt.title(r"$\frac{%s}{%s}$ over %s days" %(one,two,ndays),fontsize=12)            
    plt.tight_layout
    plt.grid()    
    n = 2
    for i in range(len(countrynames)):
        country = countrynames[i]

        filename = filenames[0]
        xdata = getdates(filename)
        ydata1 = getcountrydata(filename,country)
        filename = filenames[1]
        ydata2 = getcountrydata(filename,country)    
        theratio = []
        for j in range(len(ydata1)):
            if (ydata1[j] == 0 or ydata2[j] == 0):
                theratio.append(0)
            else:
                theratio.append(ydata1[j]/ydata2[j])      

        thenratio = getnratio(theratio)      
        plt.plot(xdata[-ndays+1::20],thenratio[-ndays+1::20])
        n = n + 1
        try:
            rmax = np.max(thenratio)
            rindex = xdata[thenratio.index(rmax)] 
            outsheet.cell(row=n,column=1).value = country
            outsheet.cell(row=n,column=2).value = rindex
        except:
            pass
        
        
        
        
        
    outputname = filetitles[0]+'over'+filetitles[1]+'.png'
    plt.show()
    plt.savefig(outputname)
    print ("Created ", outputname)
    plt.clf()
    std = outwb['Sheet']
    outwb.remove_sheet(std)
    outwb.save('maxvalues.xlsx')




def createglobal(filename,filetitle,ndays,countryname,prefix):
    import matplotlib.ticker as ticker
    import matplotlib.pyplot as plt
    print (filetitle,' for ',countryname)
    country = countryname
    ax = plt.axes()
    
    nxmajor = int(ndays/20)
    nxminor = int(ndays/10)
    ax.xaxis.set_major_locator(ticker.MultipleLocator(nxmajor))
    ax.xaxis.set_minor_locator(ticker.MultipleLocator(nxminor))
    plt.xticks(rotation=90,fontsize=10,fontweight='bold')
    plt.xlabel("Date")
    plt.ylabel(filetitle)             
    one = str(filetitle)
    plt.title(r"%s over %s days for %s " %(one,ndays,country),fontsize=12)            
    plt.tight_layout
    plt.grid()
    xdata = getdates(filename)
    ydata = getcountrydata(filename,country)
    thislabel=country
    plt.plot(xdata[-ndays+1:],ydata[-ndays+1:],label=filetitle)
    outputname = prefix+country+filetitle+'.png'
    plt.savefig(outputname)
    plt.clf()

def calculateGompertz(y):
    import math
    import numpy as np
    import sys
    print (y)
#    import time
#    time.sleep(40.0)
    
    yn = []
    yn.append(y[0])
    for i in range(1,len(y)):
        if (y[i] == 0.0 or y[i-1] == 0.0):
            print ('Problem at ', i, y[i], y[i-1])
            sys.exit()
        temp1 = y[i]/y[i-1]
        temp2 = np.log(temp1)
        if (temp1 == 1.0):
            print ("Problem - cannot take log of zero ")
            print ("The values were ", y[i], y[i-1])
            sys.exit()
        else:
            temp3 = np.log(temp2)
            yn.append(temp3)
    return yn        

def gompertz(x):
    import math
    import sys
    xn = []
    for i in range(1,len(x)):
        temp1 = x[i]/x[i-1]
        temp2 = math.log(temp1)
        if (temp2 == 0.0):
            print ("Problem encountered - cannot take log of zero ")
            sys.exit()
        else:
            xn.append(math.log(temp2))      
    return moving_average(xn,7)        

def createglobalG(filename,filetitle,ndays,countryname,prefix):
# gompertz fit     
    import matplotlib.ticker as ticker
    import matplotlib.pyplot as plt
    print (filetitle,' for ',countryname)
    country = countryname
    ax = plt.axes()
    
    nxmajor = int(ndays/20)
    nxminor = int(ndays/10)
    ax.xaxis.set_major_locator(ticker.MultipleLocator(nxmajor))
    ax.xaxis.set_minor_locator(ticker.MultipleLocator(nxminor))
    plt.xticks(rotation=90,fontsize=10,fontweight='bold')
    plt.xlabel("Date")
    plt.ylabel(filetitle)             
    one = str(filetitle)
    plt.title(r"%s over %s days for %s " %(one,ndays,country),fontsize=12)            
    plt.tight_layout
    plt.grid()
    xdata = getdates(filename)
    ydata = getcountrydata(filename,country)
    ydataG = calculateGompertz(ydata[-ndays+1:])
    thislabel=country
    plt.plot(xdata[-ndays+1:],ydataG,label=filetitle)
    outputname = prefix+country+filetitle+'.png'
    plt.savefig(outputname)
    plt.clf()


def moving_average(x, w):
    import numpy as np
#    print ("Inside mov avg")
    return np.convolve(x, np.ones(w), mode = 'same') / w
  

def createdelta(filename,filetitle,ndays,countryname,prefix):
    import matplotlib.ticker as ticker
    import matplotlib.pyplot as plt
    from matplotlib import rcParams
    rcParams.update({'figure.autolayout': True})
    rcParams['figure.dpi'] = 200
    print ("Day to day differences in", filetitle, 'for ', countryname)
    country = countryname
    ax = plt.axes()
    nxmajor = int(ndays/20)
    nxminor = int(ndays/10)
    ax.xaxis.set_major_locator(ticker.MultipleLocator(nxmajor))
    ax.xaxis.set_minor_locator(ticker.MultipleLocator(nxminor))
    plt.xticks(rotation=90,fontsize=10,fontweight='bold')
    plt.xlabel("Date")
    myylabel = "Daily difference in "+str(filetitle)
    plt.ylabel(myylabel)             
    one = str(filetitle)
    plt.title(r"Daily difference in %s over %s days for %s" %(one,ndays,country),fontsize=12)            
    plt.tight_layout
    plt.grid()
    xdata = getdates(filename)
    ydata = getcountrydata(filename,country)
    ydelta = []
    xdelta = []
    for i in range(len(ydata)-1):
        ydelta.append(ydata[i+1]-ydata[i])
        xdelta.append(xdata[i])
            
    thislabel=country
    ymovavg = moving_average(ydelta[-ndays+1:],7)
#    print (len(ymovavg), len(xdelta[-ndays+1:]))
#    plt.plot(xdelta[-ndays+1:],ydelta[-ndays+1:],label=filetitle)
    plt.plot(xdelta[-ndays+1:],ymovavg[-ndays+1:],label=filetitle)
    outputname = prefix+country+'differences'+filetitle+'.png'
    plt.savefig(outputname)
    plt.clf()


def getdelta(filename,ndays,countryname):
    xdata = getdates(filename)
    ydata = getcountrydata(filename,countryname)
    ydelta = []
    xdelta = []
    for i in range(len(ydata)-1):
        ydelta.append(ydata[i+1]-ydata[i])
        xdelta.append(xdata[i])
    ymovavg = moving_average(ydelta[-ndays+1:],7)
    return ymovavg

def createUSstateratio(filenames,filetitles,ndays,statename,prefix):
    import matplotlib.ticker as ticker
    import matplotlib.pyplot as plt
    print (filetitles[0],'/',filetitles[1], ' for ' , statename)
#    print ("Inside this one")
    ax = plt.axes()
    ax.xaxis.set_major_locator(ticker.MultipleLocator(5))
    ax.xaxis.set_minor_locator(ticker.MultipleLocator(1))
    plt.xticks(rotation=90,fontsize=10,fontweight='bold')
    plt.xlabel("Date")
    plt.ylabel("Ratio")             
    one = str(filetitles[0])
    two = str(filetitles[1])
    plt.title(r"Ratio of %s over %s for %s days for %s " %(one,two,ndays,statename),fontsize=12)            
    plt.tight_layout
    plt.grid()
    filename = filenames[0]
    xdata = getstatedates(filename)
#    print (xdata)
#    print (xdata)
    ydata1 = getstatedata(filename,statename)
#    print (ydata1)
    filename = filenames[1]
    ydata2 = getstatedata(filename,statename)    
#    print (xdata)
#    print (ydata1)
#    print (ydata2)
    theratio = []
    for i in range(len(ydata1)-1):
        if (ydata1[i] == 0 or ydata2[i] == 0):
            theratio.append(0)
        else:
            theratio.append(ydata1[i]/ydata2[i])      
    thislabel=statename
    plt.plot(xdata[-ndays+1:],theratio[-ndays+1:],label='ratio')
    outputname = prefix+filetitles[0]+filetitles[1]+'ratiosfor'+statename+'.png'
    plt.savefig(outputname)
#    print (countryname,theratio)
#    print ("Done here")
    plt.clf()


def createUSstatedelta(filename,filetitle,ndays,countryname,prefix):
    import matplotlib.ticker as ticker
    import matplotlib.pyplot as plt

    country = countryname
    ax = plt.axes()
    ax.xaxis.set_major_locator(ticker.MultipleLocator(5))
    ax.xaxis.set_minor_locator(ticker.MultipleLocator(1))
    plt.xticks(rotation=90,fontsize=10,fontweight='bold')
    plt.xlabel("Date")
    myylabel = "Daily difference in "+str(filetitle)
    plt.ylabel(myylabel)             
    one = str(filetitle)
    plt.title(r"Daily difference in %s over %s days for %s " %(one,ndays,country),fontsize=12)            
    plt.tight_layout
    plt.grid()
    filename = filename
    xdata = getstatedates(filename)
    ydata = getstatedata(filename,country)
    ydelta = []
    xdelta = []
    thedelta = []
    for i in range(len(ydata)-1):
        thedelta.append(ydata[i+1]-ydata[i])   
        xdelta.append(xdata[i])
    thislabel=country
    plt.plot(xdelta[-ndays+1:],thedelta[-ndays+1:],label=filetitle)
    outputname = prefix+filetitle+'delta'+country+'.png'
    plt.savefig(outputname)
#    print (countryname,theratio)
    print ("Created deltas for ", countryname, " in ", outputname)
    plt.clf()

def createUSstateplots(filenames,filetitles,ndays,statename,prefix):
    import matplotlib.ticker as ticker
    import matplotlib.pyplot as plt
    print ("Working on", statename)
    thisstate = statename
#    fig, ax1 = plt.subplots()
    ax1 = plt.axes()
    ax1.xaxis.set_major_locator(ticker.MultipleLocator(5))
    ax1.xaxis.set_minor_locator(ticker.MultipleLocator(1))      
    ax1.set_xlabel("Date")
    filename = filenames[0]
    xdata = getstatedates(filename)
    plt.xticks(rotation=90,fontsize=10,fontweight='bold')     
    ydata = getstatedata(filename,thisstate)
    ax1.set_ylabel(str(filetitles[0]))
    plt.tight_layout
    ax1.grid(linestyle=':')
    thislabel=filetitles[0]
    ax1.plot(xdata[-ndays+1:],ydata[-ndays+1:],label=filetitles[0])
    ax1.legend(loc=2)
    ax2 = ax1.twinx()
    filename = filenames[1]
    ydata = getstatedata(filename,thisstate)
    ax2.set_ylabel(str(filetitles[1]))
    ax2.xaxis.set_major_locator(ticker.MultipleLocator(5))
    ax2.xaxis.set_minor_locator(ticker.MultipleLocator(1))      
    
    ax2.grid(linestyle=':') 
    ax2.plot(xdata[-ndays+1:],ydata[-ndays+1:],label=filetitles[1],color='r')
    plt.title(r" %s and %s over %s days for %s" %(filetitles[0],filetitles[1],ndays,thisstate),fontsize=10)  
    ax2.legend(loc=4)
    outputname = prefix+filetitles[0]+filetitles[1]+thisstate+'.png'
    plt.savefig(outputname)
    print (" Created ", outputname)
    plt.clf()
#    fig.clf()


def twinyplots(x,xlabel,y1,y1label,c1,y2,y2label,c2,ndays,cname):
    import matplotlib.ticker as ticker
    import matplotlib.pyplot as plt
    from matplotlib import rcParams
    rcParams.update({'figure.autolayout': True})
    rcParams['figure.dpi'] = 200
#    rcParams['figure.figsize'] = 8, 4
    fig, ax1 = plt.subplots()
    plt.tight_layout
    nxmajor = int(ndays/20)
    nxminor = int(ndays/10)
    ax1.xaxis.set_major_locator(ticker.MultipleLocator(nxmajor))
    ax1.xaxis.set_minor_locator(ticker.MultipleLocator(nxminor))
    ax1.set_xlabel(xlabel)
    plt.xticks(rotation=90,fontsize=10,fontweight='bold')   
    ax1.set_ylabel(y1label,color=c1)
    plt.tight_layout
    
    ax1.grid(linestyle=':')      
    ax1.plot(x[-ndays+1:],y1[-ndays+1:],label=y1label,color=c1)
    ax1.xaxis.label.set_color('black')
    ax1.tick_params(axis='y', colors=c1)
    plt.tight_layout
    ax1.legend(loc=2)
    ax2 = ax1.twinx()
    ax2.xaxis.set_major_locator(ticker.MultipleLocator(nxmajor))
    ax2.xaxis.set_minor_locator(ticker.MultipleLocator(nxminor))      
    ax2.tick_params(axis='y', colors=c2)
    ax2.set_ylabel(y2label,color=c2)
    ax2.grid(linestyle=':') 
    ax2.plot(x[-ndays+1:],y2[-ndays+1:],label=y2label,color=c2)
    ax2.legend(loc=4)
    thistitle = str(y1label)+' and '+str(y2label)+' for '+str(cname) + ' ' + str(ndays) + ' days'
    plt.title(thistitle,fontsize=8)
#    plt.savefig(outputname)
    plt.show()
#    print (" Created ", outputname)
    plt.clf()
    fig.clf()



def createtwinyplots(x,xlabel,y1,y1label,c1,y2,y2label,c2,ndays,outputname,cname):
    import matplotlib.ticker as ticker
    import matplotlib.pyplot as plt
    from matplotlib import rcParams
    rcParams.update({'figure.autolayout': True})
    rcParams['figure.dpi'] = 200
#    rcParams['figure.figsize'] = 8, 4
    fig, ax1 = plt.subplots()
    plt.tight_layout
    nxmajor = int(ndays/20)
    nxminor = int(ndays/10)
    ax1.xaxis.set_major_locator(ticker.MultipleLocator(nxmajor))
    ax1.xaxis.set_minor_locator(ticker.MultipleLocator(nxminor))
    ax1.set_xlabel(xlabel)
    plt.xticks(rotation=90,fontsize=10,fontweight='bold')   
    ax1.set_ylabel(y1label,color=c1)
    plt.tight_layout
    
    ax1.grid(linestyle=':')      
    ax1.plot(x[-ndays+1:],y1[-ndays+1:],label=y1label,color=c1)
    ax1.xaxis.label.set_color('black')
    ax1.tick_params(axis='y', colors=c1)
    plt.tight_layout
    ax1.legend(loc=2)
    ax2 = ax1.twinx()
    ax2.xaxis.set_major_locator(ticker.MultipleLocator(nxmajor))
    ax2.xaxis.set_minor_locator(ticker.MultipleLocator(nxminor))      
    ax2.tick_params(axis='y', colors=c2)
    ax2.set_ylabel(y2label,color=c2)
    ax2.grid(linestyle=':') 
    ax2.plot(x[-ndays+1:],y2[-ndays+1:],label=y2label,color=c2)
    ax2.legend(loc=6)
    thistitle = str(y1label)+' and '+str(y2label)+' for '+str(cname)
    plt.title(thistitle,fontsize=8)
    plt.savefig(outputname)
    print (" Created ", outputname)
    plt.clf()
    fig.clf()



def createtwinyplotsG(x,xlabel,y1,y1label,c1,y2,y2label,c2,ndays,outputname,cname):
    import matplotlib.ticker as ticker
    import matplotlib.pyplot as plt
    from matplotlib import rcParams
    rcParams.update({'figure.autolayout': True})
    rcParams['figure.dpi'] = 200
#    rcParams['figure.figsize'] = 8, 4
    fig, ax1 = plt.subplots()
    plt.tight_layout
    ax1.xaxis.set_major_locator(ticker.MultipleLocator(5))
    ax1.xaxis.set_minor_locator(ticker.MultipleLocator(2))      
    ax1.set_xlabel(xlabel)
    plt.xticks(rotation=90,fontsize=10,fontweight='bold')   
    ax1.set_ylabel(y1label,color=c1)
    plt.tight_layout
    import math
    xe = []
    ye1 = []
    ye2 = []
    for i in range(1,len(y1[-ndays+1:])):
        one = y1[i+1]/y1[i]
        two = math.log(math.log(one))
        ye1.append(two)
        one = y2[i+1]/y2[i]
        two = math.log(math.log(one))
        ye2.append(two)
        xe.append(x[i])


    print (len(xe),len(ye1), len(ye2))    
    ax1.grid(linestyle=':')      
    print ('here', len(xe), len(ye1), len(ye2))
    ax1.plot(xe[-ndays+1:],ye1[-ndays+1:],label=y1label,color=c1)
    ax1.xaxis.label.set_color('black')
    ax1.tick_params(axis='y', colors=c1)
    plt.tight_layout
    ax1.legend(loc=2)
    ax2 = ax1.twinx()
    ax2.xaxis.set_major_locator(ticker.MultipleLocator(5))
    ax2.xaxis.set_minor_locator(ticker.MultipleLocator(2))      
    ax2.tick_params(axis='y', colors=c2)
    ax2.set_ylabel(y2label,color=c2)
    ax2.grid(linestyle=':') 
    ax2.plot(xe[-ndays+1:],ye2[-ndays+1:],label=y2label,color=c2)
    ax2.legend(loc=6)
    thistitle = str(y1label)+' and '+str(y2label)+' for '+str(cname)
    plt.title(thistitle,fontsize=8)
    plt.savefig(outputname)
    print (" Created ", outputname)
    plt.clf()
    fig.clf()



